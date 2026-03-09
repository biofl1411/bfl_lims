"""
엑셀 → Firestore 업로드 스크립트
바이오푸드랩_재고관리_2026년_구매처추가.xlsx → reagents 컬렉션
"""

import os
import sys
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

# ─── 설정 ───
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '바이오푸드랩_재고관리_2026년_구매처추가.xlsx')
SHEET_NAME = '2026년'
DATA_START_ROW = 5  # 4행이 헤더, 5행부터 데이터
BATCH_SIZE = 500

# Firebase Admin 초기화
SERVICE_ACCOUNT = os.path.join(os.path.dirname(__file__), '..', 'serviceAccountKey.json')
if not os.path.exists(SERVICE_ACCOUNT):
    # 서버 경로 시도
    SERVICE_ACCOUNT = os.path.expanduser('~/bfl_lims/serviceAccountKey.json')

if not firebase_admin._apps:
    cred = credentials.Certificate(SERVICE_ACCOUNT)
    firebase_admin.initialize_app(cred)

db = firestore.client()


def safe_str(val):
    """셀 값을 안전하게 문자열로 변환"""
    if val is None:
        return ''
    return str(val).strip()


def safe_int(val):
    """셀 값을 안전하게 정수로 변환 (수식 포함 처리)"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.startswith('=') or s == '' or s == '-':
        return 0
    try:
        return int(float(s.replace(',', '')))
    except (ValueError, TypeError):
        return 0


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f'엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}')
        print('프로젝트 루트에 엑셀 파일을 복사해주세요.')
        sys.exit(1)

    print(f'엑셀 파일 로드: {EXCEL_PATH}')
    wb = load_workbook(EXCEL_PATH, data_only=True)  # data_only=True: 수식 대신 계산된 값

    if SHEET_NAME not in wb.sheetnames:
        print(f'시트 "{SHEET_NAME}"를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}')
        sys.exit(1)

    ws = wb[SHEET_NAME]
    print(f'시트: {SHEET_NAME}, 전체 행 수: {ws.max_row}')

    # 데이터 수집
    reagents = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        dept = safe_str(ws.cell(row=row_num, column=1).value)       # A: 사용처
        cat = safe_str(ws.cell(row=row_num, column=2).value)        # B: 구분
        code = safe_str(ws.cell(row=row_num, column=5).value)       # E: 코드
        name = safe_str(ws.cell(row=row_num, column=6).value)       # F: 원상품명
        spec = safe_str(ws.cell(row=row_num, column=8).value)       # H: 규격
        cat_no = safe_str(ws.cell(row=row_num, column=9).value)     # I: Cat No.
        cas_no = safe_str(ws.cell(row=row_num, column=10).value)    # J: CAS.NO.
        mfr = safe_str(ws.cell(row=row_num, column=11).value)       # K: 제조사
        unit_price = safe_int(ws.cell(row=row_num, column=12).value) # L: 단가
        expiry_ref = safe_str(ws.cell(row=row_num, column=13).value) # M: 유통기한
        supplier = safe_str(ws.cell(row=row_num, column=14).value)   # N: 구매처

        # 코드가 없으면 스킵 (빈 행)
        if not code:
            continue

        barcode_val = f'BFL-{code}'

        reagents.append({
            'code': code,
            'dept': dept,
            'cat': cat,
            'name': name,
            'spec': spec,
            'catNo': cat_no,
            'casNo': cas_no,
            'mfr': mfr,
            'supplier': supplier,
            'unitPrice': unit_price,
            'expiryRef': expiry_ref,
            'stockQty': 0,
            'minQty': 1,
            'barcodeVal': barcode_val,
            'createdAt': firestore.SERVER_TIMESTAMP,
            'updatedAt': firestore.SERVER_TIMESTAMP,
        })

    print(f'변환된 데이터: {len(reagents)}건')

    if not reagents:
        print('업로드할 데이터가 없습니다.')
        sys.exit(1)

    # Firestore 배치 업로드
    total = len(reagents)
    uploaded = 0
    batch = db.batch()
    batch_count = 0

    for i, item in enumerate(reagents):
        ref = db.collection('reagents').document()  # 자동 ID
        batch.set(ref, item)
        batch_count += 1

        if batch_count >= BATCH_SIZE or i == total - 1:
            batch.commit()
            uploaded += batch_count
            print(f'  업로드: {uploaded}/{total} ({uploaded * 100 // total}%)')
            batch = db.batch()
            batch_count = 0

    print(f'\n완료! {uploaded}건 업로드됨 → Firestore "reagents" 컬렉션')


if __name__ == '__main__':
    main()
