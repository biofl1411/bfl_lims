#!/usr/bin/env python3
"""
식약처 기준규격 Excel → JSON 변환 스크립트
- 개별기준규격 (식품 시트) → individual_standards.json
- 공통기준규격 (식품 시트) → common_standards.json
- 사용여부 'Y'인 것만 추출, 날짜는 YYYY-MM-DD 문자열로 변환
"""

import openpyxl
import json
import os
import sys
from datetime import datetime

# 출력 인코딩 설정
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, '코드매핑자료')
OUTPUT_DIR = os.path.join(os.path.dirname(BASE_DIR), 'data', 'mfds')

# 출력 폴더 생성
os.makedirs(OUTPUT_DIR, exist_ok=True)


def convert_value(val):
    """셀 값을 JSON 직렬화 가능한 형태로 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()


def read_sheet(filepath, sheet_name, active_only=True):
    """Excel 시트를 dict 리스트로 변환 (식품만, 사용여부 Y만)"""
    print(f'  읽는 중: {os.path.basename(filepath)} [{sheet_name}]')
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet_name]

    # 헤더 읽기
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # 공백 제거
    headers = [h.strip() if h else h for h in headers]

    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = convert_value(val)

        # 사용여부 필터
        if active_only and row_dict.get('사용여부') != 'Y':
            skipped += 1
            continue

        records.append(row_dict)

    wb.close()
    print(f'    → {len(records)}건 추출 (스킵: {skipped}건)')
    return records


def save_json(data, filename):
    """JSON 파일 저장"""
    filepath = os.path.join(OUTPUT_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=None, separators=(',', ':'))
    size_mb = os.path.getsize(filepath) / 1024 / 1024
    print(f'  저장: {filename} ({len(data)}건, {size_mb:.1f}MB)')
    return filepath


def main():
    print('=' * 60)
    print('식약처 기준규격 Excel → JSON 변환')
    print('=' * 60)

    # 1. 개별기준규격 (식품 시트)
    print('\n[1/2] 개별기준규격 (식품)')
    ind_file = os.path.join(EXCEL_DIR, '기준_개별기준규격.xlsx')
    ind_data = read_sheet(ind_file, '식품', active_only=True)
    save_json(ind_data, 'individual_standards.json')

    # 2. 공통기준규격 (식품 시트)
    print('\n[2/2] 공통기준규격 (식품)')
    cmn_file = os.path.join(EXCEL_DIR, '기준_공통기준규격.xlsx')
    cmn_data = read_sheet(cmn_file, '식품', active_only=True)
    save_json(cmn_data, 'common_standards.json')

    # 요약
    print('\n' + '=' * 60)
    print('변환 완료!')
    print(f'  개별기준규격: {len(ind_data)}건')
    print(f'  공통기준규격: {len(cmn_data)}건')
    print(f'  합계: {len(ind_data) + len(cmn_data)}건')
    print(f'  출력 경로: {OUTPUT_DIR}')
    print('=' * 60)


if __name__ == '__main__':
    main()
